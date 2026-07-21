"""
AWS Organization Monthly Billing Report  (console-matching version)
===================================================================
Runs in the Management account (awsadmin) on the 2nd of each month.
Builds the CB Bank 13-sheet Excel report for the PREVIOUS month and
emails it via SES as an attachment.

Data source : Cost Explorer API (us-east-1)
Metric      : UnblendedCost  (same metric the AWS Console uses)

Cost basis (COST_BASIS env var)
-------------------------------
"console" (default)
    Per-account SERVICE ROWS are NET of credits/refunds (Cost Explorer
    attributes those records to the service they apply to) but GROSS of
    SPP and bundled discounts, which are broken out into their own rows
    so the arithmetic is visible on every sheet:
      * account sheet:  SubTotal + SPP + Bundled Discount = Total
        (SPP/Bundled negative here), and Total equals the console total
      * Cost+SPP+Bundle Discount / All Total: SPP and Bundled are shown
        as POSITIVE discount magnitudes, and the net total subtracts
        them:  Cost (gross) - SPP - Bundled = All Total (console net)
      * "Amazon Elastic Compute Cloud - Compute" and "EC2 - Other" are
        merged into one "Elastic Compute Cloud" line like the Bills page
      * data-transfer usage is broken out into its own "Data Transfer"
        line like the Bills page (see DATA_TRANSFER_SPLIT below)
      * service names use the console's display style ("WAF", not
        "AWS WAF")
    NOTE: because SPP/bundled are no longer netted into each service
    line, individual service rows read slightly higher than the Bills
    page; the difference sits in the SPP / Bundled Discount rows and the
    account TOTALS still match the console. Credits/Refunds remain an
    informational row (already inside the service rows).
"invoice"
    Legacy behaviour: service rows GROSS of discounts and credits
    (credits excluded entirely), SPP and bundled added below SubTotal,
    matching the invoice PDF.

Environment variables
----------------------
SENDER_EMAIL         e.g. reports@yourdomain.com   (must be SES-verified)
RECIPIENT_EMAILS     comma-separated, e.g. you@x.com,boss@x.com
REPORT_PREFIX        optional, default "CB_Bank"
COST_BASIS           optional, "console" (default) or "invoice"
ACCOUNT_SORT         optional, "cost" (default) or "name"
SES_REGION           optional, default us-east-1
DATA_TRANSFER_SPLIT  optional, "true" (default) — break data-transfer
                     usage out into its own service row like the console
DT_SOURCE_SERVICES   optional, comma-separated Cost Explorer service names
                     to pull data-transfer usage types out of. Default:
                     "EC2 - Other,Amazon Virtual Private Cloud,
                      Amazon Simple Storage Service"

IAM permissions required
------------------------
ce:GetCostAndUsage, organizations:ListAccounts, ses:SendRawEmail
"""

import os
import io
import re
import datetime
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

import boto3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Cost Explorer endpoint only exists in us-east-1
ce = boto3.client("ce", region_name="us-east-1")
org = boto3.client("organizations")
ses = boto3.client("ses", region_name=os.environ.get("SES_REGION", "us-east-1"))

# ---------------------------------------------------------------- styling
TITLE_FONT = Font(name="Arial", size=12, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True)
BODY_FONT = Font(name="Arial", size=11)
BOLD_BODY = Font(name="Arial", size=11, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '#,##0.00'


# ================================================================ dates

# ==== TESTING: set dates here manually ====
# Leave both as "" for normal auto mode (previous month).
# End date is EXCLUSIVE: for June 1-30, use end = "2026-07-01"
MANUAL_START = ""   # example: "2026-06-01"
MANUAL_END   = ""   # example: "2026-07-01"
# ==========================================


def previous_month_period(today=None):
    """Return (start_date, end_date_exclusive, label) for the previous month.
    Run on 2026-07-02 -> ('2026-06-01', '2026-07-01', 'Jun-2026')."""

    # Manual mode for testing — remember to set both back to "" after!
    if MANUAL_START and MANUAL_END:
        d = datetime.date.fromisoformat(MANUAL_START)
        return MANUAL_START, MANUAL_END, d.strftime("%b-%Y")

    today = today or datetime.date.today()
    first_this = today.replace(day=1)
    last_prev = first_this - datetime.timedelta(days=1)
    first_prev = last_prev.replace(day=1)
    label = first_prev.strftime("%b-%Y")          # Jun-2026
    return first_prev.isoformat(), first_this.isoformat(), label


# ================================================================ data
def _ce_query(start, end, group_by, flt=None):
    """get_cost_and_usage with pagination. Yields (keys_tuple, amount)."""
    kwargs = dict(
        TimePeriod={"Start": start, "End": end},
        Granularity="MONTHLY",
        Metrics=["UnblendedCost"],
        GroupBy=group_by,
    )
    if flt:
        kwargs["Filter"] = flt
    while True:
        resp = ce.get_cost_and_usage(**kwargs)
        for period in resp["ResultsByTime"]:
            for g in period["Groups"]:
                yield tuple(g["Keys"]), float(
                    g["Metrics"]["UnblendedCost"]["Amount"])
        token = resp.get("NextPageToken")
        if not token:
            break
        kwargs["NextPageToken"] = token


def _exclude_record_types_filter(record_types):
    """Not-filter for the given RECORD_TYPE values (None if empty)."""
    if not record_types:
        return None
    return {"Not": {"Dimensions": {"Key": "RECORD_TYPE",
                                   "Values": sorted(record_types)}}}


def get_account_names():
    """{account_id: name} from Organizations."""
    names = {}
    paginator = org.get_paginator("list_accounts")
    for page in paginator.paginate():
        for acct in page["Accounts"]:
            names[acct["Id"]] = acct["Name"]
    return names


def discover_record_types(start, end):
    """Which RECORD_TYPE values exist this month (Usage, Tax, Credit,
    Refund, Solution Provider Program Discount, Bundled Discount, ...)."""
    types = set()
    for (rtype,), _amount in _ce_query(
            start, end, [{"Type": "DIMENSION", "Key": "RECORD_TYPE"}]):
        types.add(rtype)
    return types


def classify_record_types(record_types):
    """Split the month's record types into the buckets the report handles
    specially. Everything NOT returned here stays in the main per-service
    data, so no record type is ever silently dropped."""
    tax = {t for t in record_types if t.lower() == "tax"}
    credit = {t for t in record_types if t.lower() in ("credit", "refund")}
    spp = {t for t in record_types
           if "solution provider" in t.lower() or "spp" in t.lower()}
    bundled = {t for t in record_types if "bundled" in t.lower()}
    return tax, credit, spp, bundled


def get_net_costs_by_account_service(start, end, exclude_types):
    """CONSOLE basis: {account_id: {service: cost}}.
    Includes usage, fees, Savings Plan records, credits and refunds —
    Cost Explorer attributes credit records to the service they apply to,
    so rows stay net of credits like the Bills page. EXCLUDES Tax (own
    row) and the SPP / bundled-discount record types, which are broken
    out into their own rows so that Cost + SPP + Bundled = Total."""
    costs = {}
    for (acct, service), amount in _ce_query(
            start, end,
            [{"Type": "DIMENSION", "Key": "LINKED_ACCOUNT"},
             {"Type": "DIMENSION", "Key": "SERVICE"}],
            _exclude_record_types_filter(exclude_types)):
        costs.setdefault(acct, {})
        costs[acct][service] = costs[acct].get(service, 0.0) + amount
    return costs


def get_costs_by_account_service(start, end, exclude_record_types):
    """INVOICE basis (legacy): {account_id: {service: gross_cost}} —
    UnblendedCost gross of credits/discounts, Tax excluded."""
    flt = _exclude_record_types_filter(exclude_record_types)
    costs = {}
    for (acct, service), amount in _ce_query(
            start, end,
            [{"Type": "DIMENSION", "Key": "LINKED_ACCOUNT"},
             {"Type": "DIMENSION", "Key": "SERVICE"}],
            flt):
        costs.setdefault(acct, {})
        costs[acct][service] = costs[acct].get(service, 0.0) + amount
    return costs


def get_special_records(start, end, tax_types, credit_types,
                        spp_types, bundled_types):
    """One query: per-account Tax, Credits/Refunds, SPP, Bundled discount.
    Returns ({acct: tax}, {acct: credits}, {acct: spp}, {acct: bundled})."""
    wanted = tax_types | credit_types | spp_types | bundled_types
    tax, credits, spp, bundled = {}, {}, {}, {}
    if not wanted:
        return tax, credits, spp, bundled

    for (acct, rtype), amount in _ce_query(
            start, end,
            [{"Type": "DIMENSION", "Key": "LINKED_ACCOUNT"},
             {"Type": "DIMENSION", "Key": "RECORD_TYPE"}],
            {"Dimensions": {"Key": "RECORD_TYPE", "Values": sorted(wanted)}}):
        if rtype in tax_types:
            tax[acct] = tax.get(acct, 0.0) + amount
        elif rtype in credit_types:
            credits[acct] = credits.get(acct, 0.0) + amount
        elif rtype in spp_types:
            spp[acct] = spp.get(acct, 0.0) + amount
        elif rtype in bundled_types:
            bundled[acct] = bundled.get(acct, 0.0) + amount
    return tax, credits, spp, bundled


# ---------------------------------------------------------- console view
# The Bills page breaks bandwidth out into its own "Data Transfer" service
# line, while Cost Explorer folds it into the owning service (mostly
# "EC2 - Other"). These usage-type markers identify the bandwidth records
# so they can be moved into a "Data Transfer" row like the console.
# CloudFront bandwidth stays under CloudFront on the Bills page, hence the
# exclusion.
DT_UT_MARKERS = ("DataTransfer", "-AWS-Out-Bytes", "-AWS-In-Bytes")


def _is_data_transfer_usage_type(usage_type):
    if "CloudFront" in usage_type:
        return False
    return any(m in usage_type for m in DT_UT_MARKERS)


def split_data_transfer(start, end, exclude_types, costs):
    """Move data-transfer usage types out of their Cost Explorer service
    and into a 'Data Transfer' row, mirroring the Bills page. One extra
    query per source service (grouped by account + usage type).

    MUST use the same record-type exclusions as the main service query —
    otherwise the amounts moved here would include record types that are
    not present in the service rows and the split would not balance."""
    if os.environ.get("DATA_TRANSFER_SPLIT", "true").lower() != "true":
        return
    sources = [s.strip() for s in os.environ.get(
        "DT_SOURCE_SERVICES",
        "EC2 - Other,Amazon Virtual Private Cloud,"
        "Amazon Simple Storage Service").split(",") if s.strip()]

    for svc in sources:
        parts = [{"Dimensions": {"Key": "SERVICE", "Values": [svc]}}]
        excl = _exclude_record_types_filter(exclude_types)
        if excl:
            parts.append(excl)
        flt = parts[0] if len(parts) == 1 else {"And": parts}

        for (acct, usage_type), amount in _ce_query(
                start, end,
                [{"Type": "DIMENSION", "Key": "LINKED_ACCOUNT"},
                 {"Type": "DIMENSION", "Key": "USAGE_TYPE"}],
                flt):
            if not _is_data_transfer_usage_type(usage_type):
                continue
            acct_costs = costs.get(acct)
            if not acct_costs or svc not in acct_costs:
                continue
            acct_costs[svc] -= amount
            acct_costs["Data Transfer"] = (
                acct_costs.get("Data Transfer", 0.0) + amount)


# The Bills page merges Cost Explorer's two EC2 entries into one line and
# drops the "Amazon"/"AWS" vendor prefix from most service names.
CONSOLE_NAME_OVERRIDES = {
    "Amazon Elastic Compute Cloud - Compute": "Elastic Compute Cloud",
    "EC2 - Other": "Elastic Compute Cloud",
    "AmazonCloudWatch": "CloudWatch",
}


def console_service_name(name):
    if name in CONSOLE_NAME_OVERRIDES:
        return CONSOLE_NAME_OVERRIDES[name]
    return re.sub(r"^(?:Amazon|AWS)\s+", "", name)


def to_console_view(costs):
    """Rename services to console display names and merge rows that map to
    the same console line (the two EC2 entries)."""
    out = {}
    for acct, svcs in costs.items():
        merged = {}
        for svc, amount in svcs.items():
            key = console_service_name(svc)
            merged[key] = merged.get(key, 0.0) + amount
        out[acct] = merged
    return out


# ================================================================ excel
def _sheet_ref(name):
    """Quote a sheet name for use inside a formula (' escaped by doubling)."""
    return "'" + name.replace("'", "''") + "'"


def _title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = TITLE_FONT


def _header_row(ws, row, col_start, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center")


def _cell(ws, row, col, value, bold=False, money=False, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = BOLD_BODY if bold else BODY_FONT
    c.border = BOX
    if money:
        c.number_format = MONEY_FMT
    if fill:
        c.fill = fill
    return c


def _merge(ws, cell_range, fill=None):
    """Style every cell in the range (border/fill) THEN merge, so the block
    renders as ONE box in Excel / Google Sheets instead of separate cells."""
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = BOX
            if fill:
                c.fill = fill
    ws.merge_cells(cell_range)


def _autofit(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_workbook(label, accounts, costs, tax, credits, spp, bundled,
                   cost_basis):
    """
    accounts : ordered list of (account_id, account_name)
    costs    : {acct: {service: cost}}  (gross of SPP/bundled; the console
               basis additionally nets credits/refunds into the rows)
    tax/credits/spp/bundled : {acct: amount}
    cost_basis : "console" or "invoice"

    Account sheet layout (both bases):
        service rows / Tax -> SubTotal
        SPP row and Bundled Discount row (negative amounts)
        Total = SubTotal + SPP + Bundled
    On the console basis that Total equals the console's account total,
    and Credits & Refunds appear below as an informational row (their
    amounts are already netted into the service rows).

    Amounts are written at FULL precision (the money format displays 2 dp)
    so sums match the console to the cent.
    """
    net_view = cost_basis == "console"
    wb = Workbook()
    wb.remove(wb.active)

    # ---- create summary sheets first so tab order matches the template
    ws_all = wb.create_sheet("All Total")
    ws_csb = wb.create_sheet("Cost+SPP+Bundle Discount")
    ws_tot = wb.create_sheet("Total Cost for All Accounts")
    ws_spp = wb.create_sheet("SPP for All Accounts")
    ws_bun = wb.create_sheet("Bundled_Discount")

    # ---- per-account sheets (dynamic service rows)
    # remember where SubTotal / SPP / Total rows land in each sheet
    anchors = {}   # acct_id -> dict(sheet, subtotal, spp, total)
    used_sheet_names = set()
    for acct_id, acct_name in accounts:
        # Strip characters Excel forbids in sheet names
        safe_name = re.sub(r"[\[\]:*?/\\]", "-", acct_name)
        sheet_name = safe_name[:31]  # Excel sheet name limit
        # Excel sheet names are case-INsensitive: "KBZMS" and "kbzms" collide,
        # and a formula ref to the loser silently reads the other sheet.
        if sheet_name.lower() in used_sheet_names:
            sheet_name = f"{safe_name[:24]}-{acct_id[-4:]}"[:31]
        ws = wb.create_sheet(sheet_name)
        # openpyxl may still rename on a collision we missed — the title it
        # actually assigned is the only name formulas can safely reference.
        sheet_name = ws.title
        used_sheet_names.add(sheet_name.lower())
        _header_row(ws, 2, 2, ["Service", "Cost"])

        services = costs.get(acct_id, {})
        row = 3
        for service in sorted(services):
            _cell(ws, row, 2, service)
            _cell(ws, row, 3, services[service], money=True)
            row += 1

        # Tax row (always present, matches template)
        _cell(ws, row, 2, "Tax")
        _cell(ws, row, 3, tax.get(acct_id, 0.0), money=True)
        row += 1

        # Service rows are GROSS of SPP/bundled on both bases now, so the
        # layout is shared: SubTotal, then the two discount rows, then
        # Total = SubTotal + SPP + Bundled. On the console basis that Total
        # equals the console's account total (credits are netted into the
        # service rows and shown below as information only).
        subtotal_row = row
        _cell(ws, row, 2, "SubTotal", bold=True)
        _cell(ws, row, 3, f"=SUM(C3:C{row - 1})", bold=True,
              money=True, fill=TOTAL_FILL)
        row += 1

        spp_row = row
        _cell(ws, row, 2, "SPP", bold=True)
        _cell(ws, row, 3, spp.get(acct_id, 0.0), money=True)
        row += 1

        bundled_row = row
        _cell(ws, row, 2, "Bundled Discount", bold=True)
        _cell(ws, row, 3, bundled.get(acct_id, 0.0), money=True)
        row += 1

        if net_view:
            _cell(ws, row, 2, "Credits & Refunds (already included above)")
            _cell(ws, row, 3, credits.get(acct_id, 0.0), money=True)
            row += 1

        total_row = row
        _cell(ws, row, 2, "Total", bold=True)
        _cell(ws, row, 3, f"=C{subtotal_row}+C{spp_row}+C{bundled_row}",
              bold=True, money=True, fill=TOTAL_FILL)

        _autofit(ws, {"B": 52, "C": 14})
        anchors[acct_id] = dict(
            sheet=sheet_name, subtotal=subtotal_row,
            spp=spp_row, bundled=bundled_row, total=total_row,
        )

    n = len(accounts)
    first, last = 4, 3 + n            # data rows 4..3+n on rollup sheets

    # ---- Total Cost for All Accounts  (pulls each sheet's SubTotal)
    _title(ws_tot, "B2", f"{label} AWS Costs for all accounts")
    _merge(ws_tot, "B2:D2")
    _header_row(ws_tot, 3, 2, ["Account Name", "Account ID", "Cost"])
    for i, (acct_id, acct_name) in enumerate(accounts):
        r = first + i
        a = anchors[acct_id]
        _cell(ws_tot, r, 2, acct_name)
        _cell(ws_tot, r, 3, acct_id)
        _cell(ws_tot, r, 4, f"={_sheet_ref(a['sheet'])}!C{a['subtotal']}",
              money=True)
    _cell(ws_tot, last + 1, 2, "Total Cost", bold=True)
    _merge(ws_tot, f"B{last + 1}:C{last + 1}")
    _cell(ws_tot, last + 1, 4, f"=SUM(D{first}:D{last})", bold=True,
          money=True, fill=TOTAL_FILL)
    _autofit(ws_tot, {"B": 22, "C": 16, "D": 14})

    # ---- SPP for All Accounts  (pulls each sheet's SPP row)
    _title(ws_spp, "B2", "Solution Provider Program Discounts for all accounts")
    _merge(ws_spp, "B2:D2")
    _header_row(ws_spp, 3, 2, ["Account Name", "Account ID", "SPP Discounts"])
    for i, (acct_id, acct_name) in enumerate(accounts):
        r = first + i
        a = anchors[acct_id]
        _cell(ws_spp, r, 2, acct_name)
        _cell(ws_spp, r, 3, acct_id)
        _cell(ws_spp, r, 4, f"={_sheet_ref(a['sheet'])}!C{a['spp']}",
              money=True)
    _cell(ws_spp, last + 1, 2, "Total Solution Provider Program Discounts",
          bold=True)
    _merge(ws_spp, f"B{last + 1}:C{last + 1}")
    _cell(ws_spp, last + 1, 4, f"=SUM(D{first}:D{last})", bold=True,
          money=True, fill=TOTAL_FILL)
    _autofit(ws_spp, {"B": 42, "C": 16, "D": 16})

    # ---- Bundled_Discount
    _title(ws_bun, "B2", "Bundled Discounts")
    _merge(ws_bun, "B2:D2")
    _header_row(ws_bun, 5, 2, ["Account Name", "Account ID", "Bundled Discount"])
    r = 6
    for acct_id, acct_name in accounts:
        amt = bundled.get(acct_id, 0.0)
        if amt:  # only list accounts that actually have bundled discounts
            _cell(ws_bun, r, 2, acct_name)
            _cell(ws_bun, r, 3, acct_id)
            _cell(ws_bun, r, 4, amt, money=True)
            r += 1
    _cell(ws_bun, r, 2, "Total Bundled Discounts", bold=True)
    _merge(ws_bun, f"B{r}:C{r}")
    if r > 6:
        _cell(ws_bun, r, 4, f"=SUM(D6:D{r - 1})", bold=True,
              money=True, fill=TOTAL_FILL)
    else:
        _cell(ws_bun, r, 4, 0, bold=True, money=True, fill=TOTAL_FILL)
    _autofit(ws_bun, {"B": 28, "C": 16, "D": 18})

    # ---- Cost+SPP+Bundle Discount
    # Columns: B=Name C=ID D=Cost E=SPP F=Bundled  (no per-row Total column)
    _title(ws_csb, "B2", f"{label} AWS Costs for all accounts")
    _merge(ws_csb, "B2:F2")
    _header_row(ws_csb, 3, 2, ["Account Name", "Account ID", "Cost",
                               "SPP Charges", "Bundled Charges"])
    for i, (acct_id, acct_name) in enumerate(accounts):
        r = first + i
        a = anchors[acct_id]
        _cell(ws_csb, r, 2, acct_name)
        _cell(ws_csb, r, 3, acct_id)
        # Cost is the gross service total. SPP and Bundled are shown here as
        # POSITIVE discount magnitudes (negated from the true negative values
        # on the account / rollup sheets), so the net Total subtracts them.
        _cell(ws_csb, r, 4, f"='Total Cost for All Accounts'!D{r}", money=True)
        _cell(ws_csb, r, 5, f"=-'SPP for All Accounts'!D{r}", money=True)
        _cell(ws_csb, r, 6, f"=-{_sheet_ref(a['sheet'])}!C{a['bundled']}",
              money=True)
    sub = last + 1
    _cell(ws_csb, sub, 2, "Sub Total", bold=True)
    _merge(ws_csb, f"B{sub}:C{sub}")
    _cell(ws_csb, sub, 4, f"=SUM(D{first}:D{last})", bold=True, money=True)
    _cell(ws_csb, sub, 5, f"=SUM(E{first}:E{last})", bold=True, money=True)
    _cell(ws_csb, sub, 6, f"=SUM(F{first}:F{last})", bold=True, money=True)
    tc = sub + 1
    # Cost is gross and SPP/Bundled are shown as positive discounts, so the
    # net Total subtracts them: Total Cost = Cost - SPP - Bundled = console
    # net = the All Total sheet.
    _cell(ws_csb, tc, 2, "Total Cost", bold=True)
    _merge(ws_csb, f"B{tc}:C{tc}")
    _cell(ws_csb, tc, 4, f"=D{sub}-E{sub}-F{sub}", bold=True, money=True)
    ws_csb.cell(row=tc, column=4).alignment = Alignment(horizontal="center")
    _merge(ws_csb, f"D{tc}:F{tc}", fill=TOTAL_FILL)
    _autofit(ws_csb, {"B": 22, "C": 16, "D": 14, "E": 14, "F": 16})

    # ---- All Total (final view)
    _title(ws_all, "B2", f"{label} AWS Costs for all accounts")
    _merge(ws_all, "B2:D2")
    _header_row(ws_all, 3, 2, ["Account Name", "Account ID", "Total Cost"])
    for i, (acct_id, acct_name) in enumerate(accounts):
        r = first + i
        # Cost - SPP - Bundled from the CSB sheet (SPP/Bundled positive there),
        # i.e. the console net cost for the account.
        formula = (f"='Cost+SPP+Bundle Discount'!D{r}"
                   f"-'Cost+SPP+Bundle Discount'!E{r}"
                   f"-'Cost+SPP+Bundle Discount'!F{r}")
        _cell(ws_all, r, 2, acct_name)
        _cell(ws_all, r, 3, acct_id)
        _cell(ws_all, r, 4, formula, money=True)
    _cell(ws_all, last + 1, 2, "Total", bold=True)
    _merge(ws_all, f"B{last + 1}:C{last + 1}")
    _cell(ws_all, last + 1, 4, f"=SUM(D{first}:D{last})", bold=True,
          money=True, fill=TOTAL_FILL)
    _autofit(ws_all, {"B": 22, "C": 16, "D": 14})

    return wb


# ================================================================ email
def send_report(wb_bytes, label, total_hint, cost_basis):
    sender = os.environ["SENDER_EMAIL"]
    recipients = [r.strip() for r in os.environ["RECIPIENT_EMAILS"].split(",")
                  if r.strip()]
    prefix = os.environ.get("REPORT_PREFIX", "CB_Bank")
    filename = f"{prefix}_{label}_AWS_Cost_Report.xlsx"

    basis_note = ("net of discounts and credits, as shown in the AWS Console"
                  if cost_basis == "console"
                  else "as invoiced, before credits")

    msg = MIMEMultipart()
    msg["Subject"] = f"AWS Cost Report - {label}"
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)

    body = (
        f"Hi,\n\n"
        f"Attached is the AWS Organization cost report for {label}.\n"
        f"Total ({basis_note}): USD {total_hint:,.2f}\n\n"
        f"Generated automatically by the billing-report Lambda.\n"
    )
    msg.attach(MIMEText(body, "plain"))

    part = MIMEApplication(wb_bytes)
    part.add_header("Content-Disposition", "attachment", filename=filename)
    part.add_header(
        "Content-Type",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    msg.attach(part)

    ses.send_raw_email(
        Source=sender,
        Destinations=recipients,
        RawMessage={"Data": msg.as_string()},
    )
    logger.info("Report %s sent to %s", filename, recipients)
    return filename


# ================================================================ handler
def lambda_handler(event, context):
    # Optional overrides for backfills / testing:
    #   {"start": "2026-06-01", "end": "2026-07-01", "cost_basis": "console"}
    event = event or {}
    if event.get("start") and event.get("end"):
        start, end = event["start"], event["end"]
        d = datetime.date.fromisoformat(start)
        label = d.strftime("%b-%Y")
    else:
        start, end, label = previous_month_period()

    cost_basis = (event.get("cost_basis")
                  or os.environ.get("COST_BASIS", "console")).lower()
    if cost_basis not in ("console", "invoice"):
        raise ValueError(f"COST_BASIS must be 'console' or 'invoice', "
                         f"got {cost_basis!r}")

    logger.info("Reporting period %s -> %s (%s), basis=%s",
                start, end, label, cost_basis)

    names = get_account_names()
    record_types = discover_record_types(start, end)
    logger.info("Record types present: %s", record_types)

    tax_types, credit_types, spp_types, bundled_types = \
        classify_record_types(record_types)

    tax, credits, spp, bundled = get_special_records(
        start, end, tax_types, credit_types, spp_types, bundled_types)

    if cost_basis == "console":
        # Service rows net of credits/refunds but GROSS of SPP and bundled
        # discounts, which get their own rows/columns so that
        # Cost + SPP + Bundled = Total on every sheet. The data-transfer
        # split must use the exact same exclusions to stay in balance.
        exclude = tax_types | spp_types | bundled_types
        costs = get_net_costs_by_account_service(start, end, exclude)
        split_data_transfer(start, end, exclude, costs)
        costs = to_console_view(costs)
    else:
        # Legacy invoice basis: gross of credits/discounts.
        costs = get_costs_by_account_service(
            start, end,
            exclude_record_types=(tax_types | credit_types
                                  | spp_types | bundled_types))

    # Account list = every account in the Organization, plus any account that
    # appears in cost data but is no longer in the org (e.g. removed mid-month).
    all_ids = (set(names) | set(costs) | set(tax) | set(credits)
               | set(spp) | set(bundled))

    def month_total(a):
        # Service rows are gross of SPP/bundled on both bases now, so the
        # discounts are added back here. Credits are netted into the rows
        # on the console basis and excluded on the invoice basis, so the
        # console value still equals the console's account total.
        return (sum(costs.get(a, {}).values()) + tax.get(a, 0.0)
                + spp.get(a, 0.0) + bundled.get(a, 0.0))

    if os.environ.get("ACCOUNT_SORT", "cost").lower() == "name":
        ordered = sorted(all_ids, key=lambda a: names.get(a, a).lower())
    else:  # cost: highest spender first
        ordered = sorted(all_ids, key=month_total, reverse=True)

    accounts = [(a, names.get(a, a)) for a in ordered]

    wb = build_workbook(label, accounts, costs, tax, credits, spp, bundled,
                        cost_basis)

    buf = io.BytesIO()
    wb.save(buf)
    wb_bytes = buf.getvalue()

    total = sum(month_total(a) for a in all_ids)
    credits_total = sum(credits.values())
    if cost_basis == "console":
        console_total = total
        invoice_total = total - credits_total
    else:
        invoice_total = total
        console_total = total + credits_total

    filename = send_report(wb_bytes, label, total, cost_basis)
    return {
        "status": "sent",
        "file": filename,
        "period": f"{start}..{end}",
        "cost_basis": cost_basis,
        "total": round(total, 2),
        "console_total": round(console_total, 2),   # matches Bills page
        "invoice_total": round(invoice_total, 2),   # matches invoice PDF
        "credits_refunds": round(credits_total, 2),
        "spp_total": round(sum(spp.values()), 2),
        "bundled_total": round(sum(bundled.values()), 2),
    }
