"""
AWS Organizations monthly billing report Lambda.

Design:
- No CUR.
- No S3.
- Workbook is generated in /tmp and attached directly through Amazon SES.
- openpyxl must be supplied by a Lambda layer.
- Account worksheets contain only: Service | Cost.
- Account service costs use Cost Explorer SERVICE + USAGE_TYPE with
  NetUnblendedCost by default. Marketplace products keep their Cost Explorer
  product-service label. To match the Bills view, DataTransfer/DataXfer usage
  rows from any service, plus EC2/ELB AWS in/out byte rows, are moved into
  Data Transfer in the same query.

Required environment variables:
- SES_SENDER
- SES_RECIPIENTS                 Comma-separated addresses

Recommended environment variables:
- SES_REGION                     Default: AWS_REGION or us-east-1
- SES_CC                         Optional comma-separated CC addresses
- COST_EXPLORER_REGION           Default: us-east-1
- ORGANIZATIONS_REGION           Default: us-east-1
- COMPONENT_METRIC               Default: UnblendedCost
- SERVICE_METRIC                 Default: NetUnblendedCost
- CURRENCY_SYMBOL                Default: $
- SPP_RECORD_TYPES               Default: Solution Provider Program Discount
- BUNDLED_RECORD_TYPES           Default: Bundled Discount,BundledDiscount
- RECLASSIFY_DATA_TRANSFER       Default: true
- DATA_TRANSFER_USAGE_PATTERNS   Default: DataTransfer,DataXfer
- SERVICE_NAME_MAP_JSON          Optional JSON object for custom name overrides
- FAIL_ON_ESTIMATED              Default: true
- FAIL_ON_RECONCILIATION         Default: true
- RECONCILIATION_TOLERANCE       Absolute floor in currency units. Default: 0.02
- RECONCILIATION_REL_TOLERANCE   Fraction of the account total. Default: 0.005
- SES_RAW_EMAIL_MAX_BYTES        Default: 9500000
- LOG_LEVEL                      Default: INFO

Event flags:
- report_month      "YYYY-MM"; defaults to the previous calendar month
- send_email        Default: true. NOTE: an empty test event {} sends the email.
- force_estimated   Default: false. Allows a provisional report while Cost
                    Explorer still marks the period as Estimated.

Test events:
    {"report_month": "2026-06", "send_email": true}
    {"send_email": false}

Cost Explorer End date is exclusive. For June 2026 the query period is:
Start=2026-06-01, End=2026-07-01.

Reconciliation: the account total (COMPONENT_METRIC, UnblendedCost by default)
is compared against the sum of the account's displayed service costs
(SERVICE_METRIC, NetUnblendedCost by default). Because these are different
metrics, discount netting can legitimately diverge by a spend-proportional
amount, so the effective tolerance per account is
max(RECONCILIATION_TOLERANCE, RECONCILIATION_REL_TOLERANCE * account total).
"""

from __future__ import annotations

import json
import logging
import os
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from email import policy
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import boto3
from botocore.config import Config
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LOGGER = logging.getLogger(__name__)
ZERO = Decimal("0")

BOTO_CONFIG = Config(
    retries={"max_attempts": 10, "mode": "adaptive"},
    connect_timeout=10,
    read_timeout=120,
)

# SendRawEmail is not idempotent: a retry after a read timeout can deliver
# the report twice, so the SES client must not retry automatically.
SES_BOTO_CONFIG = Config(
    retries={"max_attempts": 1},
    connect_timeout=10,
    read_timeout=120,
)

SES_ABSOLUTE_MAX_BYTES = 10_000_000
SES_DEFAULT_SAFE_MAX_BYTES = 9_500_000


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return ZERO
    try:
        result = Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value!r}") from exc
    if not result.is_finite():
        raise ValueError(f"Non-finite decimal value: {value!r}")
    return result


def decimal_env(name: str, default: str) -> Decimal:
    raw = os.getenv(name, default).strip() or default
    try:
        return decimal_value(raw)
    except ValueError as exc:
        raise ValueError(f"{name} must be a finite decimal number, got: {raw!r}") from exc


def normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def parse_csv_env(name: str, default: Sequence[str]) -> tuple[str, ...]:
    raw = os.getenv(name)
    source = default if not raw else raw.split(",")
    return tuple(item.strip() for item in source if item.strip())


def parse_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def parse_bool_env(name: str, default: bool) -> bool:
    return parse_bool(os.getenv(name), default)


def required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise ValueError(f"Missing required environment variable: {name}")
    return value


def split_addresses(raw: str) -> list[str]:
    return [item.strip() for item in raw.replace(";", ",").split(",") if item.strip()]


def previous_month_period(today: date | None = None) -> tuple[str, str, str]:
    today = today or datetime.now(timezone.utc).date()
    current_month_start = date(today.year, today.month, 1)
    if current_month_start.month == 1:
        previous_start = date(current_month_start.year - 1, 12, 1)
    else:
        previous_start = date(
            current_month_start.year,
            current_month_start.month - 1,
            1,
        )
    return (
        previous_start.isoformat(),
        current_month_start.isoformat(),
        previous_start.strftime("%b-%Y"),
    )


def period_from_month(month: str) -> tuple[str, str, str]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", month.strip())
    if not match:
        raise ValueError("report_month must use YYYY-MM format")

    year = int(match.group(1))
    month_number = int(match.group(2))
    if not 1 <= month_number <= 12:
        raise ValueError("report_month month must be between 01 and 12")

    start = date(year, month_number, 1)
    end = date(year + 1, 1, 1) if month_number == 12 else date(year, month_number + 1, 1)
    return start.isoformat(), end.isoformat(), start.strftime("%b-%Y")


# ---------------------------------------------------------------------------
# Configuration and data models
# ---------------------------------------------------------------------------


DEFAULT_SERVICE_NAME_MAP: dict[str, str] = {
    "Amazon Athena": "Athena",
    "Amazon Elastic Compute Cloud - Compute": "Elastic Compute Cloud",
    "EC2 - Other": "Elastic Compute Cloud",
    "Amazon Elastic Load Balancing": "Elastic Load Balancing",
    "Elastic Load Balancing": "Elastic Load Balancing",
    "AmazonCloudWatch": "CloudWatch",
    "AWS CloudTrail": "CloudTrail",
    "AWS Cost Explorer": "Cost Explorer",
    "AWS Glue": "Glue",
    "AWS Key Management Service": "Key Management Service",
    "AWS WAF": "WAF",
    "Amazon Simple Notification Service": "Simple Notification Service",
    "Amazon Simple Storage Service": "Simple Storage Service",
    "Amazon Virtual Private Cloud": "Virtual Private Cloud",
    "AWS Data Transfer": "Data Transfer",
    "Data Transfer": "Data Transfer",
    "Savings Plans for AWS Compute usage": "Savings Plans for AWS Compute usage",
}


@dataclass(frozen=True)
class ReportConfig:
    component_metric: str
    service_metric: str
    currency_symbol: str
    tolerance: Decimal
    relative_tolerance: Decimal
    fail_on_estimated: bool
    fail_on_reconciliation: bool
    reclassify_data_transfer: bool
    data_transfer_usage_patterns: tuple[str, ...]
    spp_record_types: tuple[str, ...]
    bundled_record_types: tuple[str, ...]
    credit_record_types: tuple[str, ...]
    refund_record_types: tuple[str, ...]
    tax_record_types: tuple[str, ...]
    savings_plan_record_types: tuple[str, ...]
    generic_discount_record_types: tuple[str, ...]
    service_name_map: Mapping[str, str]

    @classmethod
    def from_env(cls) -> "ReportConfig":
        custom_map_raw = os.getenv("SERVICE_NAME_MAP_JSON", "").strip()
        custom_map: dict[str, str] = {}
        if custom_map_raw:
            try:
                parsed = json.loads(custom_map_raw)
            except json.JSONDecodeError as exc:
                raise ValueError("SERVICE_NAME_MAP_JSON must be valid JSON") from exc
            if not isinstance(parsed, dict):
                raise ValueError("SERVICE_NAME_MAP_JSON must be a JSON object")
            custom_map = {str(key): str(value) for key, value in parsed.items()}

        merged_map = dict(DEFAULT_SERVICE_NAME_MAP)
        merged_map.update(custom_map)

        return cls(
            component_metric=os.getenv("COMPONENT_METRIC", "UnblendedCost"),
            service_metric=os.getenv("SERVICE_METRIC", "NetUnblendedCost"),
            currency_symbol=os.getenv("CURRENCY_SYMBOL", "$"),
            tolerance=decimal_env("RECONCILIATION_TOLERANCE", "0.02"),
            relative_tolerance=decimal_env("RECONCILIATION_REL_TOLERANCE", "0.005"),
            fail_on_estimated=parse_bool_env("FAIL_ON_ESTIMATED", True),
            fail_on_reconciliation=parse_bool_env("FAIL_ON_RECONCILIATION", True),
            reclassify_data_transfer=parse_bool_env("RECLASSIFY_DATA_TRANSFER", True),
            data_transfer_usage_patterns=parse_csv_env(
                "DATA_TRANSFER_USAGE_PATTERNS",
                ("DataTransfer", "DataXfer"),
            ),
            spp_record_types=parse_csv_env(
                "SPP_RECORD_TYPES",
                ("Solution Provider Program Discount",),
            ),
            bundled_record_types=parse_csv_env(
                "BUNDLED_RECORD_TYPES",
                ("Bundled Discount", "BundledDiscount"),
            ),
            credit_record_types=parse_csv_env("CREDIT_RECORD_TYPES", ("Credit",)),
            refund_record_types=parse_csv_env("REFUND_RECORD_TYPES", ("Refund",)),
            tax_record_types=parse_csv_env("TAX_RECORD_TYPES", ("Tax",)),
            savings_plan_record_types=parse_csv_env(
                "SAVINGS_PLAN_RECORD_TYPES",
                (
                    "SavingsPlanCoveredUsage",
                    "Savings Plan Covered Usage",
                    "SavingsPlanNegation",
                    "Savings Plan Negation",
                    "SavingsPlanRecurringFee",
                    "Savings Plan Recurring Fee",
                    "SavingsPlanUpfrontFee",
                    "Savings Plan Upfront Fee",
                ),
            ),
            generic_discount_record_types=parse_csv_env(
                "GENERIC_DISCOUNT_RECORD_TYPES",
                ("Discount",),
            ),
            service_name_map=merged_map,
        )


@dataclass(frozen=True)
class Account:
    account_id: str
    name: str


@dataclass
class AccountBreakdown:
    account: Account
    direct_total: Decimal = ZERO
    unit: str = "USD"
    record_types: dict[str, Decimal] = field(default_factory=dict)
    display_services: dict[str, Decimal] = field(default_factory=dict)
    raw_service_usage: dict[tuple[str, str], Decimal] = field(default_factory=dict)


@dataclass
class BillingDataset:
    start_date: str
    end_date: str
    month_label: str
    accounts: list[Account]
    account_breakdowns: dict[str, AccountBreakdown]
    estimated: bool
    warnings: list[str]


class RecordTypeClassifier:
    def __init__(self, config: ReportConfig):
        self.config = config
        self.groups = {
            "spp": {normalize(item) for item in config.spp_record_types},
            "bundled": {normalize(item) for item in config.bundled_record_types},
            "credit": {normalize(item) for item in config.credit_record_types},
            "refund": {normalize(item) for item in config.refund_record_types},
            "tax": {normalize(item) for item in config.tax_record_types},
            "savings_plans": {
                normalize(item) for item in config.savings_plan_record_types
            },
            "other_discount": {
                normalize(item) for item in config.generic_discount_record_types
            },
        }

    def category(self, record_type: str) -> str:
        key = normalize(record_type)
        for category, values in self.groups.items():
            if key in values:
                return category
        return "base"


# ---------------------------------------------------------------------------
# Cost Explorer and Organizations collection
# ---------------------------------------------------------------------------


class BillingCollector:
    def __init__(
        self,
        ce_client: Any,
        organizations_client: Any,
        config: ReportConfig,
    ):
        self.ce = ce_client
        self.organizations = organizations_client
        self.config = config
        self._normalized_service_map = {
            normalize(key): value for key, value in config.service_name_map.items()
        }
        self._transfer_patterns = tuple(
            normalize(item) for item in config.data_transfer_usage_patterns
        )

    def _cost_pages(self, request: Mapping[str, Any]) -> Iterable[dict[str, Any]]:
        next_token: str | None = None
        while True:
            page_request = dict(request)
            if next_token:
                page_request["NextPageToken"] = next_token
            response = self.ce.get_cost_and_usage(**page_request)
            yield response
            next_token = response.get("NextPageToken")
            if not next_token:
                break

    def list_active_accounts(self) -> list[Account]:
        accounts: list[Account] = []
        next_token: str | None = None
        while True:
            request: dict[str, Any] = {"MaxResults": 20}
            if next_token:
                request["NextToken"] = next_token
            response = self.organizations.list_accounts(**request)
            for item in response.get("Accounts", []):
                state = item.get("State") or item.get("Status") or "UNKNOWN"
                if state == "ACTIVE":
                    accounts.append(
                        Account(
                            account_id=item["Id"],
                            name=item.get("Name") or item["Id"],
                        )
                    )
            next_token = response.get("NextToken")
            if not next_token:
                break
        return sorted(accounts, key=lambda item: (item.name.lower(), item.account_id))

    def discover_record_types(self, start_date: str, end_date: str) -> list[str]:
        values: set[str] = set()
        next_token: str | None = None
        while True:
            request: dict[str, Any] = {
                "TimePeriod": {"Start": start_date, "End": end_date},
                "Dimension": "RECORD_TYPE",
                "Context": "COST_AND_USAGE",
                "MaxResults": 1000,
            }
            if next_token:
                request["NextPageToken"] = next_token
            response = self.ce.get_dimension_values(**request)
            for item in response.get("DimensionValues", []):
                if item.get("Value"):
                    values.add(item["Value"])
            next_token = response.get("NextPageToken")
            if not next_token:
                break
        return sorted(values, key=str.lower)

    def get_account_record_types(
        self,
        start_date: str,
        end_date: str,
    ) -> tuple[dict[str, dict[str, Decimal]], dict[str, Decimal], dict[str, str], bool]:
        by_account: dict[str, dict[str, Decimal]] = defaultdict(
            lambda: defaultdict(lambda: ZERO)
        )
        organization_totals: dict[str, Decimal] = defaultdict(lambda: ZERO)
        units: dict[str, str] = {}
        estimated = False
        request = {
            "TimePeriod": {"Start": start_date, "End": end_date},
            "Granularity": "MONTHLY",
            "Metrics": [self.config.component_metric],
            "GroupBy": [
                {"Type": "DIMENSION", "Key": "LINKED_ACCOUNT"},
                {"Type": "DIMENSION", "Key": "RECORD_TYPE"},
            ],
        }
        for response in self._cost_pages(request):
            for period in response.get("ResultsByTime", []):
                estimated = estimated or bool(period.get("Estimated"))
                for group in period.get("Groups", []):
                    keys = group.get("Keys", [])
                    if len(keys) != 2:
                        continue
                    account_id, record_type = keys
                    metric = group.get("Metrics", {}).get(
                        self.config.component_metric,
                        {},
                    )
                    amount = decimal_value(metric.get("Amount"))
                    by_account[account_id][record_type] += amount
                    organization_totals[record_type] += amount
                    if metric.get("Unit"):
                        units[account_id] = metric["Unit"]
        return (
            {account_id: dict(values) for account_id, values in by_account.items()},
            dict(organization_totals),
            units,
            estimated,
        )

    def _display_service_name(self, raw_service: str, usage_type: str) -> str:
        raw_normalized = normalize(raw_service)
        if self.config.reclassify_data_transfer:
            return self._transfer_target(raw_service, usage_type) or self._normalized_service_map.get(raw_normalized, raw_service)
        return self._normalized_service_map.get(raw_normalized, raw_service)

    def _transfer_target(self, raw_service: str, usage_type: str) -> str | None:
        """Return Data Transfer only for the CE rows verified against Bills.

        The Bills view assigns DataTransfer/DataXfer usage to Data Transfer,
        including the matching VPC rows. EC2 AWS in/out byte rows are the
        additional transfer class needed for this account.
        """
        service = normalize(raw_service)
        usage = normalize(usage_type)
        if service not in {normalize("Data Transfer"), normalize("AWS Data Transfer")} and any(
            pattern and pattern in usage for pattern in self._transfer_patterns
        ):
            return "Data Transfer"
        if service in {
            normalize("Amazon Elastic Compute Cloud - Compute"),
            normalize("EC2 - Other"),
            normalize("Amazon Elastic Load Balancing"),
            normalize("Elastic Load Balancing"),
        } and any(pattern in usage for pattern in ("awsinbytes", "awsoutbytes")):
            return "Data Transfer"
        return None

    def get_display_services_for_account(
        self,
        account_id: str,
        start_date: str,
        end_date: str,
    ) -> tuple[dict[str, Decimal], dict[tuple[str, str], Decimal], bool]:
        display_services: dict[str, Decimal] = defaultdict(lambda: ZERO)
        raw_service_usage: dict[tuple[str, str], Decimal] = defaultdict(lambda: ZERO)
        estimated = False
        request = {
            "TimePeriod": {"Start": start_date, "End": end_date},
            "Granularity": "MONTHLY",
            "Metrics": [self.config.service_metric],
            "Filter": {
                "Dimensions": {
                    "Key": "LINKED_ACCOUNT",
                    "Values": [account_id],
                }
            },
            "GroupBy": [
                {"Type": "DIMENSION", "Key": "SERVICE"},
                {"Type": "DIMENSION", "Key": "USAGE_TYPE"},
            ],
        }
        for response in self._cost_pages(request):
            for period in response.get("ResultsByTime", []):
                estimated = estimated or bool(period.get("Estimated"))
                for group in period.get("Groups", []):
                    keys = group.get("Keys", [])
                    if len(keys) != 2:
                        continue
                    raw_service, usage_type = keys
                    metric = group.get("Metrics", {}).get(
                        self.config.service_metric,
                        {},
                    )
                    amount = decimal_value(metric.get("Amount"))
                    raw_service_usage[(raw_service, usage_type)] += amount
                    display_name = self._display_service_name(raw_service, usage_type)
                    display_services[display_name] += amount

        return dict(display_services), dict(raw_service_usage), estimated

    def collect(
        self,
        start_date: str,
        end_date: str,
        month_label: str,
    ) -> BillingDataset:
        accounts = self.list_active_accounts()
        account_map = {account.account_id: account for account in accounts}

        discovered_record_types = self.discover_record_types(start_date, end_date)
        (
            account_record_types,
            organization_record_type_totals,
            units,
            estimated_record_types,
        ) = self.get_account_record_types(start_date, end_date)

        # The account total is the sum of its record-type amounts; a separate
        # LINKED_ACCOUNT-only query would return the same numbers.
        direct_totals = {
            account_id: sum(values.values(), ZERO)
            for account_id, values in account_record_types.items()
        }

        for account_id in sorted(account_record_types):
            if account_id not in account_map:
                fallback = Account(account_id=account_id, name=f"Account {account_id}")
                accounts.append(fallback)
                account_map[account_id] = fallback

        breakdowns: dict[str, AccountBreakdown] = {}
        estimated = estimated_record_types

        for account in sorted(accounts, key=lambda item: (item.name.lower(), item.account_id)):
            services, raw_service_usage, service_estimated = (
                self.get_display_services_for_account(
                    account.account_id,
                    start_date,
                    end_date,
                )
            )
            estimated = estimated or service_estimated
            breakdowns[account.account_id] = AccountBreakdown(
                account=account,
                direct_total=direct_totals.get(account.account_id, ZERO),
                unit=units.get(account.account_id, "USD"),
                record_types=account_record_types.get(account.account_id, {}),
                display_services=services,
                raw_service_usage=raw_service_usage,
            )

        classifier = RecordTypeClassifier(self.config)
        warnings: list[str] = []
        discovered_normalized = {normalize(item) for item in discovered_record_types}
        discovered_display = ", ".join(discovered_record_types) or "none"

        if not any(
            normalize(item) in discovered_normalized
            for item in self.config.spp_record_types
        ):
            warnings.append(
                "No configured SPP RECORD_TYPE was discovered for this period. "
                f"Cost Explorer returned these record types: {discovered_display}."
            )
        if not any(
            normalize(item) in discovered_normalized
            for item in self.config.bundled_record_types
        ):
            warnings.append(
                "No configured bundled-discount RECORD_TYPE was discovered for this period. "
                f"Cost Explorer returned these record types: {discovered_display}."
            )

        generic_discount_total = sum(
            (
                amount
                for record_type, amount in organization_record_type_totals.items()
                if classifier.category(record_type) == "other_discount"
            ),
            ZERO,
        )
        if generic_discount_total != ZERO:
            warnings.append(
                "A generic Discount RECORD_TYPE has a non-zero total of "
                f"{generic_discount_total}; it is not broken out into its own "
                "column in the workbook."
            )

        return BillingDataset(
            start_date=start_date,
            end_date=end_date,
            month_label=month_label,
            accounts=sorted(accounts, key=lambda item: (item.name.lower(), item.account_id)),
            account_breakdowns=breakdowns,
            estimated=estimated,
            warnings=warnings,
        )


# ---------------------------------------------------------------------------
# Excel workbook generation
# ---------------------------------------------------------------------------


class ExcelReportBuilder:
    HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
    TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
    THIN_GRAY = Side(style="thin", color="B7B7B7")
    TOP_BORDER = Border(top=Side(style="thin", color="000000"))

    def __init__(self, config: ReportConfig):
        self.config = config
        self.classifier = RecordTypeClassifier(config)
        self.currency_format = (
            f'{config.currency_symbol}#,##0.00;[Red]('
            f'{config.currency_symbol}#,##0.00);-'
        )

    @staticmethod
    def _sheet_name(workbook: Workbook, desired: str) -> str:
        base = re.sub(r"[\\/*?:\[\]]", "-", desired).strip() or "Account"
        base = base[:31]
        candidate = base
        counter = 2
        while candidate in workbook.sheetnames:
            suffix = f"-{counter}"
            candidate = f"{base[:31-len(suffix)]}{suffix}"
            counter += 1
        return candidate

    @staticmethod
    def _set_widths(worksheet: Any, widths: Mapping[int, float]) -> None:
        for column, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column)].width = width

    def _title(self, worksheet: Any, text: str, end_column: int) -> None:
        worksheet.merge_cells(
            start_row=2,
            start_column=2,
            end_row=2,
            end_column=end_column,
        )
        cell = worksheet.cell(2, 2, text)
        cell.font = Font(name="Calibri", size=16, bold=True)
        cell.alignment = Alignment(horizontal="left")

    def _headers(
        self,
        worksheet: Any,
        row: int,
        start_column: int,
        headers: Sequence[str],
    ) -> None:
        for offset, header in enumerate(headers):
            cell = worksheet.cell(row, start_column + offset, header)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=self.THIN_GRAY)

    def _currency_cells(
        self,
        worksheet: Any,
        rows: Iterable[int],
        columns: Iterable[int],
    ) -> None:
        for row in rows:
            for column in columns:
                worksheet.cell(row, column).number_format = self.currency_format
                worksheet.cell(row, column).alignment = Alignment(horizontal="right")

    def _components(self, breakdown: AccountBreakdown) -> dict[str, Decimal]:
        values = {
            "base": ZERO,
            "savings_plans": ZERO,
            "spp": ZERO,
            "bundled": ZERO,
            "credit": ZERO,
            "refund": ZERO,
            "tax": ZERO,
            "other_discount": ZERO,
        }
        for record_type, amount in breakdown.record_types.items():
            values[self.classifier.category(record_type)] += amount
        return values

    def _new_sheet(self, workbook: Workbook, name: str) -> Any:
        return workbook.create_sheet(name)

    def _write_all_total(self, workbook: Workbook, dataset: BillingDataset) -> None:
        worksheet = self._new_sheet(workbook, "All Total")
        self._title(
            worksheet,
            f"{dataset.month_label} AWS Costs for all accounts",
            4,
        )
        self._headers(
            worksheet,
            3,
            2,
            ["Account Name", "Account ID", "Total Cost"],
        )

        row = 4
        for account in dataset.accounts:
            breakdown = dataset.account_breakdowns[account.account_id]
            worksheet.cell(row, 2, account.name)
            worksheet.cell(row, 3, account.account_id)
            worksheet.cell(row, 4, float(breakdown.direct_total))
            row += 1

        worksheet.cell(row, 2, "Total").font = Font(bold=True)
        total_formula = f"=SUM(D4:D{row - 1})" if row > 4 else 0
        worksheet.cell(row, 4, total_formula).font = Font(bold=True)
        for column in range(2, 5):
            worksheet.cell(row, column).fill = self.TOTAL_FILL
            worksheet.cell(row, column).border = self.TOP_BORDER

        self._currency_cells(worksheet, range(4, row + 1), [4])
        self._set_widths(worksheet, {2: 36, 3: 18, 4: 20})
        worksheet.freeze_panes = "B4"
        worksheet.auto_filter.ref = f"B3:D{max(3, row - 1)}"

    def _write_components(self, workbook: Workbook, dataset: BillingDataset) -> None:
        worksheet = self._new_sheet(workbook, "Cost+SPP+Bundle Discount")
        self._title(worksheet, f"{dataset.month_label} AWS Costs for all accounts", 6)
        headers = [
            "Account Name",
            "Account ID",
            "Cost",
            "SPP Charges",
            "Bundled Charges",
        ]
        self._headers(worksheet, 3, 2, headers)

        row = 4
        for account in dataset.accounts:
            breakdown = dataset.account_breakdowns[account.account_id]
            components = self._components(breakdown)
            values: list[Any] = [
                account.name,
                account.account_id,
                components["base"],
                components["spp"],
                components["bundled"],
            ]
            for column, value in enumerate(values, 2):
                worksheet.cell(
                    row,
                    column,
                    float(value) if isinstance(value, Decimal) else value,
                )
            row += 1

        worksheet.cell(row, 2, "Sub Total").font = Font(bold=True)
        for column in range(4, 7):
            letter = get_column_letter(column)
            formula = f"=SUM({letter}4:{letter}{row - 1})" if row > 4 else 0
            worksheet.cell(row, column, formula).font = Font(bold=True)
        for column in range(2, 7):
            worksheet.cell(row, column).fill = self.TOTAL_FILL
            worksheet.cell(row, column).border = self.TOP_BORDER

        # Same figure as the "All Total" sheet's Total row, referenced directly
        # so the two sheets can never disagree. That cell sits one row below
        # the per-account rows, which both sheets draw from dataset.accounts.
        all_total_row = 4 + len(dataset.accounts)
        row += 1
        worksheet.cell(row, 2, "Total Cost").font = Font(bold=True)
        worksheet.cell(row, 4, f"='All Total'!D{all_total_row}").font = Font(bold=True)
        for column in range(2, 7):
            worksheet.cell(row, column).fill = self.TOTAL_FILL
            worksheet.cell(row, column).border = self.TOP_BORDER

        self._currency_cells(worksheet, range(4, row + 1), range(4, 7))
        self._set_widths(
            worksheet,
            {2: 36, 3: 18, **{column: 18 for column in range(4, 7)}},
        )
        worksheet.freeze_panes = "B4"
        worksheet.auto_filter.ref = f"B3:F{max(3, row - 2)}"

    def _write_account_category_summary(
        self,
        workbook: Workbook,
        dataset: BillingDataset,
        sheet_name: str,
        title: str,
        category: str,
        amount_header: str,
    ) -> None:
        worksheet = self._new_sheet(workbook, sheet_name)
        headers = ["Account Name", "Account ID", amount_header]
        end_column = 2 + len(headers) - 1
        self._title(worksheet, title, end_column)
        self._headers(worksheet, 3, 2, headers)

        row = 4
        for account in dataset.accounts:
            breakdown = dataset.account_breakdowns[account.account_id]
            amount = self._components(breakdown)[category]
            worksheet.cell(row, 2, account.name)
            worksheet.cell(row, 3, account.account_id)
            worksheet.cell(row, 4, float(amount))
            row += 1

        worksheet.cell(row, 2, "Total").font = Font(bold=True)
        total_formula = f"=SUM(D4:D{row - 1})" if row > 4 else 0
        worksheet.cell(row, 4, total_formula).font = Font(bold=True)
        last_column = 4
        for column in range(2, last_column + 1):
            worksheet.cell(row, column).fill = self.TOTAL_FILL
            worksheet.cell(row, column).border = self.TOP_BORDER

        self._currency_cells(worksheet, range(4, row + 1), [4])
        self._set_widths(
            worksheet,
            {2: 36, 3: 18, 4: 22},
        )
        worksheet.freeze_panes = "B4"
        worksheet.auto_filter.ref = (
            f"B3:{get_column_letter(last_column)}{max(3, row - 1)}"
        )

    def _reconciliation_errors(self, dataset: BillingDataset) -> list[str]:
        """Keep the Lambda safety check without adding a worksheet.

        The account total (COMPONENT_METRIC) and the displayed service sum
        (SERVICE_METRIC) are different Cost Explorer metrics, so discount
        netting can legitimately diverge in proportion to spend. The effective
        tolerance is therefore the greater of the absolute floor and the
        spend-relative allowance.
        """
        errors: list[str] = []
        for account in dataset.accounts:
            breakdown = dataset.account_breakdowns[account.account_id]
            service_diff = breakdown.direct_total - sum(
                breakdown.display_services.values(), ZERO
            )
            effective_tolerance = max(
                self.config.tolerance,
                abs(breakdown.direct_total) * self.config.relative_tolerance,
            )
            if abs(service_diff) > effective_tolerance:
                errors.append(
                    f"{account.name} ({account.account_id}): "
                    f"account total={breakdown.direct_total}, "
                    f"displayed service sum diff={service_diff}, "
                    f"tolerance={effective_tolerance}"
                )
        return errors

    def _write_account_sheets(self, workbook: Workbook, dataset: BillingDataset) -> None:
        """Each account sheet intentionally contains only Service and Cost."""
        for account in dataset.accounts:
            breakdown = dataset.account_breakdowns[account.account_id]
            worksheet = workbook.create_sheet(
                self._sheet_name(
                    workbook,
                    f"{account.name}-{account.account_id[-4:]}",
                )
            )
            self._title(
                worksheet,
                f"{dataset.month_label} - {account.name} ({account.account_id})",
                3,
            )
            self._headers(worksheet, 3, 2, ["Service", "Cost"])

            row = 4
            for service, amount in sorted(
                breakdown.display_services.items(),
                key=lambda item: item[0].lower(),
            ):
                display_amount = ZERO if abs(amount) < Decimal("0.005") else amount
                worksheet.cell(row, 2, service)
                worksheet.cell(row, 3, float(display_amount))
                row += 1

            if row == 4:
                worksheet.cell(row, 2, "No service costs returned for this period.")
                worksheet.cell(row, 3, 0)
                row += 1

            worksheet.cell(row, 2, "Total").font = Font(bold=True)
            worksheet.cell(row, 3, f"=SUM(C4:C{row - 1})").font = Font(bold=True)
            for column in range(2, 4):
                worksheet.cell(row, column).fill = self.TOTAL_FILL
                worksheet.cell(row, column).border = self.TOP_BORDER

            self._currency_cells(worksheet, range(4, row + 1), [3])
            self._set_widths(worksheet, {2: 52, 3: 22})
            worksheet.freeze_panes = "B4"
            worksheet.auto_filter.ref = f"B3:C{max(3, row - 1)}"

    def build(self, dataset: BillingDataset, output_path: str) -> list[str]:
        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_all_total(workbook, dataset)
        self._write_components(workbook, dataset)
        self._write_account_category_summary(
            workbook,
            dataset,
            "Total Cost for All Accounts",
            f"{dataset.month_label} AWS Costs for all accounts",
            "base",
            "Cost",
        )
        self._write_account_category_summary(
            workbook,
            dataset,
            "SPP for All Accounts",
            f"{dataset.month_label} Solution Provider Program Discounts",
            "spp",
            "SPP Discounts",
        )
        self._write_account_category_summary(
            workbook,
            dataset,
            "Bundled_Discount",
            f"{dataset.month_label} Bundled Discounts",
            "bundled",
            "Bundled Discount",
        )
        reconciliation_errors = self._reconciliation_errors(dataset)
        self._write_account_sheets(workbook, dataset)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return reconciliation_errors


# ---------------------------------------------------------------------------
# Direct SES delivery
# ---------------------------------------------------------------------------


def send_report_email(
    ses_client: Any,
    sender: str,
    recipients: list[str],
    cc: list[str],
    subject: str,
    body_html: str,
    report_path: str,
) -> dict[str, Any]:
    if not recipients:
        raise ValueError("At least one SES recipient is required")

    all_recipients = recipients + cc
    if len(all_recipients) > 50:
        raise ValueError("SES SendRawEmail supports at most 50 total recipients")

    try:
        configured_max = int(
            os.getenv("SES_RAW_EMAIL_MAX_BYTES", str(SES_DEFAULT_SAFE_MAX_BYTES))
        )
    except ValueError as exc:
        raise ValueError("SES_RAW_EMAIL_MAX_BYTES must be an integer") from exc

    if configured_max <= 0:
        raise ValueError("SES_RAW_EMAIL_MAX_BYTES must be greater than zero")
    max_raw_bytes = min(configured_max, SES_ABSOLUTE_MAX_BYTES)

    report = Path(report_path)
    if not report.exists():
        raise FileNotFoundError(f"Workbook does not exist: {report_path}")

    message = MIMEMultipart("mixed")
    message["Subject"] = subject
    message["From"] = sender
    message["To"] = ", ".join(recipients)
    if cc:
        message["Cc"] = ", ".join(cc)

    alternative = MIMEMultipart("alternative")
    alternative.attach(
        MIMEText(
            "The AWS monthly billing workbook is attached.",
            "plain",
            "utf-8",
        )
    )
    alternative.attach(MIMEText(body_html, "html", "utf-8"))
    message.attach(alternative)

    with report.open("rb") as report_file:
        attachment = MIMEApplication(
            report_file.read(),
            _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    attachment.add_header("Content-Disposition", "attachment", filename=report.name)
    message.attach(attachment)

    raw_message = message.as_bytes(policy=policy.SMTP)
    raw_size = len(raw_message)
    if raw_size > max_raw_bytes:
        raise ValueError(
            f"SES email is too large: {raw_size:,} bytes; "
            f"configured limit is {max_raw_bytes:,} bytes."
        )

    response = ses_client.send_raw_email(
        Source=sender,
        Destinations=all_recipients,
        RawMessage={"Data": raw_message},
    )
    return {
        "message_id": response.get("MessageId"),
        "raw_message_bytes": raw_size,
        "attachment_bytes": report.stat().st_size,
    }


# ---------------------------------------------------------------------------
# Lambda entry point
# ---------------------------------------------------------------------------


def lambda_handler(event: dict[str, Any] | None, context: Any) -> dict[str, Any]:
    event = event or {}

    log_level_name = os.getenv("LOG_LEVEL", "INFO").upper()
    LOGGER.setLevel(getattr(logging, log_level_name, logging.INFO))

    config = ReportConfig.from_env()

    report_month = event.get("report_month")
    if report_month:
        start_date, end_date, month_label = period_from_month(str(report_month))
    else:
        start_date, end_date, month_label = previous_month_period()

    ce_client = boto3.client(
        "ce",
        region_name=os.getenv("COST_EXPLORER_REGION", "us-east-1"),
        config=BOTO_CONFIG,
    )
    organizations_client = boto3.client(
        "organizations",
        region_name=os.getenv("ORGANIZATIONS_REGION", "us-east-1"),
        config=BOTO_CONFIG,
    )
    ses_client = boto3.client(
        "ses",
        region_name=os.getenv("SES_REGION", os.getenv("AWS_REGION", "us-east-1")),
        config=SES_BOTO_CONFIG,
    )

    sender = required_env("SES_SENDER")
    recipients = split_addresses(required_env("SES_RECIPIENTS"))
    cc = split_addresses(os.getenv("SES_CC", ""))

    output_name = f"{month_label}_AWS_Cost_Report.xlsx"
    output_path = Path("/tmp") / output_name

    try:
        collector = BillingCollector(
            ce_client=ce_client,
            organizations_client=organizations_client,
            config=config,
        )
        dataset = collector.collect(start_date, end_date, month_label)

        force_estimated = parse_bool(event.get("force_estimated"), False)
        if dataset.estimated and config.fail_on_estimated and not force_estimated:
            raise RuntimeError(
                "Cost Explorer marked this period as Estimated. "
                "Wait for finalized data or set force_estimated=true for a provisional report."
            )

        builder = ExcelReportBuilder(config)
        reconciliation_errors = builder.build(dataset, str(output_path))

        if reconciliation_errors and config.fail_on_reconciliation:
            raise RuntimeError(
                "Reconciliation failed; email was not sent: "
                + " | ".join(reconciliation_errors[:10])
            )

        send_email = parse_bool(event.get("send_email"), True)
        email_result: dict[str, Any] = {
            "message_id": None,
            "raw_message_bytes": None,
            "attachment_bytes": output_path.stat().st_size,
        }

        if send_email:
            warning_prefix = "[WARNING] " if dataset.warnings else ""
            email_result = send_report_email(
                ses_client=ses_client,
                sender=sender,
                recipients=recipients,
                cc=cc,
                subject=f"{warning_prefix}{month_label} AWS Monthly Billing Report",
                body_html=f"""
                <html><body>
                  <p>The AWS monthly billing workbook for
                     <strong>{month_label}</strong> is attached.</p>
                  <p>Billing period: {start_date} through {end_date}
                     (end date exclusive).</p>
                  <p>Accounts included: {len(dataset.accounts)}</p>
                  <p>Estimated data: {'Yes' if dataset.estimated else 'No'}</p>
                  <p>Warnings: {len(dataset.warnings)}</p>
                  <p>No CUR and no S3 storage were used.</p>
                </body></html>
                """,
                report_path=str(output_path),
            )
        else:
            LOGGER.warning(
                "send_email=false: workbook was generated in /tmp and will now be deleted"
            )

        result = {
            "status": "ok",
            "billing_period": {"start": start_date, "end": end_date},
            "month_label": month_label,
            "accounts": len(dataset.accounts),
            "estimated": dataset.estimated,
            "warnings": dataset.warnings,
            "reconciliation_errors": reconciliation_errors,
            "report_filename": output_name,
            "attachment_bytes": email_result["attachment_bytes"],
            "raw_message_bytes": email_result["raw_message_bytes"],
            "email_sent": send_email,
            "ses_message_id": email_result["message_id"],
            "persistent_storage": False,
        }
        LOGGER.info("Billing report result: %s", json.dumps(result, default=str))
        return result
    finally:
        try:
            output_path.unlink(missing_ok=True)
        except OSError:
            LOGGER.exception("Could not delete temporary workbook: %s", output_path)
